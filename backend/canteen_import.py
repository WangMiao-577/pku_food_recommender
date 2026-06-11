"""
canteen_import.py - 从食堂 Excel 导入菜品（农园等）
首列文字为详细档口位置，写入 location_hint / window / floor。
"""

import json
import re
from pathlib import Path
from typing import Dict, List, Optional, Tuple

CANTEEN_META = {
    "农园食堂": {"canteen_id": "nongyuan", "cuisine_default": "融合"},
    "燕南美食": {"canteen_id": "yannan", "cuisine_default": "融合"},
    "家园食堂": {"canteen_id": "jiayuan", "cuisine_default": "融合"},
    "畅春园": {"canteen_id": "changchun", "cuisine_default": "融合"},
    "佟园": {"canteen_id": "tongyuan", "cuisine_default": "清真"},
    "松林": {"canteen_id": "songlin", "cuisine_default": "融合"},
    "学一食堂": {"canteen_id": "xueyi", "cuisine_default": "融合"},
    "勺园": {"canteen_id": "shaoyuan", "cuisine_default": "融合"},
}

# 导航/展示时 location_hint 仅保留食堂名
SIMPLE_LOCATION_CANTEENS = {"佟园", "松林"}

# 导入时统一归并到的食堂名（源 JSON 中的别名 -> 标准名）
CANTEEN_ALIASES = {
    "雁南地下": "燕南美食",
    "燕南地下": "燕南美食",
    "家园三楼": "家园食堂",
    "家园三层": "家园食堂",
}

REPLACE_CANTEENS = {"燕南美食", "雁南地下", "燕南地下", "家园食堂", "家园三楼", "家园三层"}

WINDOW_CUISINE = {
    "日韩料理": "日韩", "西北风味": "西北", "西式简餐": "西式",
    "清真": "清真", "家常菜": "融合", "粤菜": "粤",
    "东侧窗口": "湘", "西侧窗口": "湘", "水饺": "融合",
    "水果": "轻食", "铁板炒饭": "融合", "麻辣烫": "川",
    "蜀湘风味": "湘", "广东风味": "粤", "西北风味档": "西北",
    "家常小炒": "融合", "面食档口": "融合", "家园素食": "融合",
    "风味面食": "融合", "北侧 主食窗口": "融合", "南侧 风味窗口": "融合",
}


def parse_calories(raw) -> float:
    if raw is None:
        return 400.0
    s = str(raw).strip()
    m = re.search(r"([\d.]+)\s*kcal", s, re.I)
    if m:
        return float(m.group(1))
    m = re.search(r"([\d.]+)\s*千焦", s)
    if m:
        return round(float(m.group(1)) / 4.184, 1)
    if isinstance(raw, (int, float)):
        return float(raw)
    m = re.search(r"([\d.]+)", s)
    return float(m.group(1)) if m else 400.0


def parse_gram_value(raw) -> float:
    if raw is None or raw == "":
        return 0.0
    m = re.search(r"([\d.]+)", str(raw))
    return float(m.group(1)) if m else 0.0


def parse_flavor_text(text: str) -> List[str]:
    text = (text or "").strip()
    if not text:
        return ["鲜"]
    tags = []
    rules = [
        ("麻辣", "麻辣"), ("微辣", "微辣"), ("鲜辣", "辣"), ("香辣", "辣"),
        ("酸辣", "酸"), ("酸甜", "甜"), ("咸香", "咸"), ("清淡", "清淡"),
        ("甜", "甜"), ("酸", "酸"), ("辣", "辣"), ("鲜", "鲜"), ("咸", "咸"),
        ("糯", "鲜"), ("嫩", "鲜"), ("香", "鲜"),
    ]
    for kw, tag in rules:
        if kw in text and tag not in tags:
            tags.append(tag)
    return tags[:3] or ["鲜"]


def parse_csv_location(raw: str, default_prefix: str) -> Tuple[str, int, str]:
    """解析 CSV 首列档口：如「家园二层 家园素食」「燕南地上 北侧 主食窗口」"""
    loc = (raw or "").strip()
    if not loc:
        loc = default_prefix

    floor = 1
    if "地下" in loc:
        floor = 0
    elif "二层" in loc or "二楼" in loc:
        floor = 2
    elif "三层" in loc or "三楼" in loc:
        floor = 3
    elif "地上" in loc:
        floor = 1

    window = "综合"
    for sep in ("，", ","):
        if sep in loc:
            window = loc.split(sep, 1)[1].strip()
            break
    else:
        parts = loc.split()
        if len(parts) >= 2:
            window = " ".join(parts[1:]) if " " in loc else parts[-1]
        elif parts:
            window = parts[0]

    hint = loc if "，" in loc or "," in loc else loc.replace("  ", "，").replace(" ", "，", 1) if " " in loc else loc
    if "，" not in hint and default_prefix and default_prefix not in hint:
        hint = f"{default_prefix}，{window}"

    return window, floor, hint


def infer_price_from_context(name: str, window: str, calories: float) -> float:
    if any(k in window for k in ("称重", "自选")):
        return 15.0
    if any(k in name for k in ("粥", "馒头", "窝头", "小菜")):
        return 3.0
    if calories >= 600 or any(k in name for k in ("套餐", "大碗", "排骨")):
        return 18.0
    if any(k in window for k in ("素食", "小炒")):
        return 10.0
    if any(k in name for k in ("面", "粉", "饭")):
        return 14.0
    if any(k in window for k in ("烘焙", "点心", "主食")):
        return 8.0
    return 12.0


def infer_cuisine(window: str, canteen: str, flavor_text: str = "") -> str:
    for key, cuisine in WINDOW_CUISINE.items():
        if key in window:
            return cuisine
    text = window + flavor_text
    if any(k in text for k in ("湘", "辣")):
        return "湘"
    if "粤" in text or "广" in text:
        return "粤"
    if "川" in text or "麻辣" in text:
        return "川"
    if "西北" in text or "拉面" in text:
        return "西北"
    return CANTEEN_META.get(canteen, {}).get("cuisine_default", "融合")


def parse_price(raw) -> float:
    if raw is None:
        return 0.0
    s = str(raw).replace("元", "").strip()
    m = re.search(r"[\d.]+", s)
    return float(m.group()) if m else 0.0


def parse_location_hint(loc: str) -> Tuple[str, int, str]:
    """解析「农园二层，东侧窗口」-> floor, window, full hint"""
    loc = (loc or "").strip()
    floor = 1
    if "地下" in loc:
        floor = 0
    elif "二层" in loc or "二楼" in loc:
        floor = 2
    elif "三层" in loc or "三楼" in loc:
        floor = 3

    window = "综合"
    if "，" in loc:
        window = loc.split("，", 1)[1].strip()
    elif "," in loc:
        window = loc.split(",", 1)[1].strip()

    return window, floor, loc


def infer_flavor(name: str, ingredients: str) -> List[str]:
    text = name + ingredients
    flavors = []
    rules = [
        ("麻辣", "麻辣"), ("微辣", "微辣"), ("辣", "辣"), ("酸", "酸"),
        ("甜", "甜"), ("鲜", "鲜"), ("咸", "咸"), ("清淡", "清淡"),
    ]
    for kw, tag in rules:
        if kw in text and tag not in flavors:
            flavors.append(tag)
    if not flavors:
        if any(k in text for k in ("粉", "面", "饭")):
            flavors = ["咸", "鲜"]
        else:
            flavors = ["鲜"]
    return flavors[:3]


def infer_nutrition(calories: float) -> Dict[str, float]:
    cal = float(calories or 400)
    protein = round(cal * 0.12 / 4, 1)
    fat = round(cal * 0.28 / 9, 1)
    carbs = round(cal * 0.45 / 4, 1)
    return {"calories": int(cal), "protein": protein, "fat": fat, "carbs": carbs, "fiber": 2.0}


def infer_tags(name: str, window: str, price: float) -> List[str]:
    tags = []
    if price <= 10:
        tags.append("经济")
    if any(k in name for k in ("粉", "面", "饭", "饺")):
        tags.append("管饱")
    if "水果" in window:
        tags.append("健康")
    if any(k in name for k in ("牛肉", "鸡", "排骨", "肉")):
        tags.append("高蛋白")
    if "水饺" in window or "饺" in name:
        tags.append("经典")
    return tags or ["家常"]


def row_to_dish(row: Dict, dish_id: str, canteen: str, window_numbers: Dict[str, int]) -> Dict:
    window, floor, hint = parse_location_hint(row["location"])
    if window not in window_numbers:
        window_numbers[window] = len(window_numbers) + 1

    name = row["name"]
    ingredients = row.get("ingredients", "")
    price = parse_price(row.get("price_raw"))
    nutrition = infer_nutrition(row.get("calories", 400))
    cuisine = WINDOW_CUISINE.get(window, CANTEEN_META[canteen]["cuisine_default"])

    return {
        "id": dish_id,
        "name": name,
        "canteen": canteen,
        "canteen_id": CANTEEN_META[canteen]["canteen_id"],
        "window": window,
        "window_number": window_numbers[window],
        "floor": floor,
        "price": price,
        "cuisine": cuisine,
        "flavor": infer_flavor(name, ingredients),
        "cooking": "现做",
        "appearance": 3,
        "portion_size": "M" if price < 14 else "L",
        "prep_time": 5 if "水果" in window else 8,
        "image": "",
        "tags": infer_tags(name, window, price),
        "rating": 4.0,
        "rating_count": 0,
        "hours": {"lunch": True, "dinner": True, "late_night": False},
        "related_dishes": [],
        "location_hint": hint,
        "ingredients": ingredients,
        **nutrition,
    }


def load_nongyuan_from_excel(xlsx_path: str) -> List[Dict]:
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = []
    loc = ""
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(c for c in row if c):
            continue
        if row[1]:
            loc = str(row[1]).strip()
        name = row[2] if len(row) > 2 else None
        if not name or str(name).strip() in ("菜品名称",):
            continue
        rows.append({
            "location": loc,
            "name": str(name).strip(),
            "price_raw": row[3] if len(row) > 3 else "",
            "ingredients": str(row[4] or "") if len(row) > 4 else "",
            "calories": row[5] if len(row) > 5 else 400,
        })
    return rows


def merge_nongyuan_dishes(
    dishes_file: str,
    xlsx_path: str,
    id_prefix: str = "ny",
    start_index: int = 1,
) -> Tuple[int, int]:
    """导入农园菜品并合并到 dishes.json，返回 (新增, 跳过)"""
    path = Path(dishes_file)
    existing = json.loads(path.read_text(encoding="utf-8")) if path.exists() else []
    existing_names = {(d["canteen"], d["name"]) for d in existing}
    existing_ids = {d["id"] for d in existing}

    raw_rows = load_nongyuan_from_excel(xlsx_path)
    window_numbers: Dict[str, int] = {}
    added = 0
    skipped = 0
    idx = start_index

    for row in raw_rows:
        if row["location"] in ("图片来源", "") or not row["name"]:
            skipped += 1
            continue
        key = ("农园食堂", row["name"])
        if key in existing_names:
            skipped += 1
            continue
        dish_id = f"{id_prefix}{idx:03d}"
        while dish_id in existing_ids:
            idx += 1
            dish_id = f"{id_prefix}{idx:03d}"
        dish = row_to_dish(row, dish_id, "农园食堂", window_numbers)
        existing.append(dish)
        existing_names.add(key)
        existing_ids.add(dish_id)
        added += 1
        idx += 1

    path.write_text(json.dumps(existing, ensure_ascii=False, indent=2), encoding="utf-8")
    return added, skipped


def _default_price(price, window: str) -> float:
    if price is not None and price != "":
        return float(price)
    if "称重" in window:
        return 15.0
    return 0.0


def normalize_imported_dish(
    raw: Dict,
    dish_id: str,
    canteen: str,
    floor: int,
    location_prefix: str,
    window_numbers: Dict[str, int],
) -> Dict:
    """将外部 JSON 菜品规范为系统 dishes.json 结构"""
    window = (raw.get("window") or "综合").strip()
    if window not in window_numbers:
        window_numbers[window] = len(window_numbers) + 1

    meta = CANTEEN_META.get(canteen, {"canteen_id": canteen, "cuisine_default": "融合"})
    price = _default_price(raw.get("price"), window)
    hint = f"{location_prefix}，{window}" if location_prefix else window

    dish = {
        "id": dish_id,
        "name": raw["name"],
        "canteen": canteen,
        "canteen_id": meta["canteen_id"],
        "window": window,
        "window_number": window_numbers[window],
        "floor": floor,
        "price": price,
        "cuisine": raw.get("cuisine") or meta["cuisine_default"],
        "flavor": raw.get("flavor") or ["鲜"],
        "cooking": raw.get("cooking") or "现做",
        "appearance": raw.get("appearance", 3),
        "portion_size": raw.get("portion_size") or ("M" if price < 14 else "L"),
        "prep_time": raw.get("prep_time", 8),
        "image": raw.get("image") or "",
        "tags": raw.get("tags") or ["家常"],
        "rating": raw.get("rating", 4.0),
        "rating_count": raw.get("rating_count", 0),
        "hours": raw.get("hours") or {"lunch": True, "dinner": True, "late_night": False},
        "related_dishes": raw.get("related_dishes") or [],
        "location_hint": hint,
        "calories": int(raw.get("calories") or 400),
        "protein": raw.get("protein", 0),
        "fat": raw.get("fat", 0),
        "carbs": raw.get("carbs", 0),
        "fiber": raw.get("fiber", 2),
    }
    if price == 15.0 and raw.get("price") is None and "称重" in window:
        if "称重" not in dish["tags"]:
            dish["tags"] = list(dish["tags"]) + ["称重"]
    return dish


def merge_json_canteen_file(
    dishes_file: str,
    json_path: str,
    *,
    target_canteen: str,
    floor: int,
    location_prefix: str,
    id_prefix: str,
    replace_existing: bool = True,
) -> Tuple[int, int]:
    """
    从食堂 JSON 文件导入并合并到 dishes.json。
    replace_existing=True 时先移除该食堂已有菜品再写入。
    返回 (新增, 跳过重复名)
    """
    path = Path(dishes_file)
    existing = json.loads(path.read_text(encoding="utf-8")) if path.exists() else []

    if replace_existing:
        remove_names = {target_canteen}
        for alias, canonical in CANTEEN_ALIASES.items():
            if canonical == target_canteen:
                remove_names.add(alias)
        existing = [d for d in existing if d.get("canteen") not in remove_names]

    raw_list = json.loads(Path(json_path).read_text(encoding="utf-8"))
    existing_names = {(d["canteen"], d["name"]) for d in existing}
    existing_ids = {d["id"] for d in existing}
    window_numbers: Dict[str, int] = {}
    added = 0
    skipped = 0
    idx = 1

    for raw in raw_list:
        name = (raw.get("name") or "").strip()
        if not name:
            skipped += 1
            continue
        if (target_canteen, name) in existing_names:
            skipped += 1
            continue

        dish_id = f"{id_prefix}{idx:03d}"
        while dish_id in existing_ids:
            idx += 1
            dish_id = f"{id_prefix}{idx:03d}"

        dish = normalize_imported_dish(
            raw, dish_id, target_canteen, floor, location_prefix, window_numbers
        )
        existing.append(dish)
        existing_names.add((target_canteen, name))
        existing_ids.add(dish_id)
        added += 1
        idx += 1

    path.write_text(json.dumps(existing, ensure_ascii=False, indent=2), encoding="utf-8")
    return added, skipped


def _next_dish_id(prefix: str, idx: int, existing_ids: set) -> Tuple[str, int]:
    dish_id = f"{prefix}{idx:03d}"
    while dish_id in existing_ids:
        idx += 1
        dish_id = f"{prefix}{idx:03d}"
    return dish_id, idx


def load_nutrition_csv(csv_path: str, encoding: str = "gbk") -> List[Dict]:
    """家园/燕南格式：图片编号, 菜品名称, 主要原料, 口味特点, 营养成分, 热量"""
    import csv

    rows = []
    current_loc = ""
    with open(csv_path, encoding=encoding, newline="") as f:
        reader = csv.reader(f)
        header = next(reader, None)
        for row in reader:
            if not row or not any(str(c).strip() for c in row):
                continue
            if row[0].strip():
                current_loc = row[0].strip()
            name = row[1].strip() if len(row) > 1 else ""
            if not name or name in ("菜品名称",):
                continue
            rows.append({
                "location": current_loc,
                "name": name,
                "ingredients": row[2].strip() if len(row) > 2 else "",
                "flavor_text": row[3].strip() if len(row) > 3 else "",
                "nutrition_text": row[4].strip() if len(row) > 4 else "",
                "calories_raw": row[5].strip() if len(row) > 5 else "",
            })
    return rows


def load_changchun_csv(csv_path: str, encoding: str = "gbk") -> List[Dict]:
    """畅春园格式：地点, 菜品名称, 原料, 口味, 蛋白质…热量"""
    import csv

    rows = []
    current_loc = ""
    with open(csv_path, encoding=encoding, newline="") as f:
        reader = csv.reader(f)
        next(reader, None)
        for row in reader:
            if not row or not any(str(c).strip() for c in row):
                continue
            if row[0].strip():
                current_loc = row[0].strip()
            name = row[1].strip() if len(row) > 1 else ""
            if not name:
                continue
            rows.append({
                "location": current_loc,
                "name": name,
                "ingredients": row[2].strip() if len(row) > 2 else "",
                "flavor_text": row[3].strip() if len(row) > 3 else "",
                "protein": parse_gram_value(row[4] if len(row) > 4 else ""),
                "fat": parse_gram_value(row[5] if len(row) > 5 else ""),
                "carbs": parse_gram_value(row[6] if len(row) > 6 else ""),
                "fiber": parse_gram_value(row[7] if len(row) > 7 else ""),
                "sodium": parse_gram_value(row[8] if len(row) > 8 else ""),
                "calcium": parse_gram_value(row[9] if len(row) > 9 else ""),
                "iron": parse_gram_value(row[10] if len(row) > 10 else ""),
                "vitamin_c": parse_gram_value(row[11] if len(row) > 11 else ""),
                "calories": parse_calories(row[12] if len(row) > 12 else ""),
            })
    return rows


def csv_row_to_dish(
    row: Dict,
    dish_id: str,
    canteen: str,
    location_prefix: str,
    window_numbers: Dict[str, int],
    *,
    floor_override: int = None,
) -> Dict:
    window, floor, hint = parse_csv_location(row.get("location", ""), location_prefix)
    if floor_override is not None:
        floor = floor_override

    if window not in window_numbers:
        window_numbers[window] = len(window_numbers) + 1

    name = row["name"]
    ingredients = row.get("ingredients", "")
    flavor_text = row.get("flavor_text", "")
    calories = row.get("calories") or parse_calories(row.get("calories_raw", 400))

    if row.get("protein"):
        nutrition = {
            "calories": int(calories),
            "protein": row.get("protein", 0),
            "fat": row.get("fat", 0),
            "carbs": row.get("carbs", 0),
            "fiber": row.get("fiber", 2),
        }
    else:
        nutrition = infer_nutrition(calories)

    price = infer_price_from_context(name, window, calories)
    cuisine = infer_cuisine(window, canteen, flavor_text)
    meta = CANTEEN_META.get(canteen, {"canteen_id": canteen, "cuisine_default": "融合"})

    dish = {
        "id": dish_id,
        "name": name,
        "canteen": canteen,
        "canteen_id": meta["canteen_id"],
        "window": window,
        "window_number": window_numbers[window],
        "floor": floor,
        "price": price,
        "cuisine": cuisine,
        "flavor": parse_flavor_text(flavor_text) or infer_flavor(name, ingredients),
        "cooking": "现做",
        "appearance": 3,
        "portion_size": "M" if calories < 350 else "L",
        "prep_time": 8,
        "image": "",
        "tags": infer_tags(name, window, price),
        "rating": 4.0,
        "rating_count": 0,
        "hours": {"lunch": True, "dinner": True, "late_night": "夜宵" in window or "风味" in window},
        "related_dishes": [],
        "location_hint": hint,
        "ingredients": ingredients,
        **nutrition,
    }
    if row.get("nutrition_text"):
        dish["nutrition_notes"] = row["nutrition_text"]
    if row.get("sodium"):
        dish["sodium_mg"] = row.get("sodium")
    return dish


def merge_csv_canteen(
    dishes_file: str,
    csv_path: str,
    *,
    target_canteen: str,
    location_prefix: str,
    id_prefix: str,
    loader: str = "nutrition",
    encoding: str = "gbk",
    floor_override: int = None,
    replace_existing: bool = False,
) -> Tuple[int, int]:
    """
    从 CSV 导入菜品并合并到 dishes.json。
    loader: nutrition（家园/燕南）| changchun
    replace_existing=True 时先移除该食堂已有菜品。
    """
    path = Path(dishes_file)
    existing = json.loads(path.read_text(encoding="utf-8")) if path.exists() else []

    if replace_existing:
        existing = [d for d in existing if d.get("canteen") != target_canteen]

    if loader == "changchun":
        raw_rows = load_changchun_csv(csv_path, encoding=encoding)
    else:
        raw_rows = load_nutrition_csv(csv_path, encoding=encoding)

    existing_names = {(d["canteen"], d["name"]) for d in existing}
    existing_ids = {d["id"] for d in existing}
    window_numbers: Dict[str, int] = {}
    added = 0
    skipped = 0
    idx = 1

    for row in raw_rows:
        name = (row.get("name") or "").strip()
        if not name:
            skipped += 1
            continue
        if (target_canteen, name) in existing_names:
            skipped += 1
            continue

        dish_id, idx = _next_dish_id(id_prefix, idx, existing_ids)
        dish = csv_row_to_dish(
            row, dish_id, target_canteen, location_prefix, window_numbers,
            floor_override=floor_override,
        )
        existing.append(dish)
        existing_names.add((target_canteen, name))
        existing_ids.add(dish_id)
        added += 1
        idx += 1

    path.write_text(json.dumps(existing, ensure_ascii=False, indent=2), encoding="utf-8")
    return added, skipped


def _simple_location_hint(canteen: str, window: str) -> str:
    if canteen in SIMPLE_LOCATION_CANTEENS:
        return canteen
    return window


def parse_tongyuan_window(loc: str) -> str:
    loc = (loc or "").strip()
    if "风味小吃" in loc or "风味小食" in loc:
        return "风味小吃"
    if "清真" in loc:
        return "清真档口"
    if "夜宵" in loc:
        return "夜宵"
    if "午餐" in loc:
        return "午餐"
    return "民族餐厅"


def parse_songlin_window(loc: str) -> str:
    loc = (loc or "").strip()
    for key in ("包子类", "小食类", "粥类", "饮品", "灯箱"):
        if key in loc:
            return key.replace("类", "")
    if "燕园味道" in loc:
        return "饮品"
    return "快餐"


def resolve_xueyi_shaoyuan_canteen(loc: str, last_canteen: str) -> str:
    loc = (loc or "").strip()
    if not loc or loc.startswith("图片"):
        return last_canteen or "学一食堂"
    if loc.startswith("勺") or "勺园" in loc:
        return "勺园"
    if "学一" in loc or "学医" in loc:
        return "学一食堂"
    return last_canteen or "学一食堂"


def load_tongyuan_csv(csv_path: str, encoding: str = "utf-8-sig") -> List[Dict]:
    """佟园：含营养数据与「周几有」列，合并同名同档口的多日记录"""
    import csv
    from backend.dish_availability import parse_weekday_schedule

    merged: Dict[Tuple[str, str], Dict] = {}
    current_loc = ""
    with open(csv_path, encoding=encoding, newline="") as f:
        reader = csv.reader(f)
        next(reader, None)
        for row in reader:
            if not row or not any(str(c).strip() for c in row):
                continue
            if row[0].strip():
                current_loc = row[0].strip()
            name = row[1].strip() if len(row) > 1 else ""
            if not name:
                continue
            day_text = row[-1].strip() if row else ""
            key = (name, current_loc)
            parsed_days = parse_weekday_schedule(day_text)
            if key not in merged:
                merged[key] = {
                    "location": current_loc,
                    "name": name,
                    "ingredients": row[2].strip() if len(row) > 2 else "",
                    "flavor_text": row[3].strip() if len(row) > 3 else "",
                    "protein": parse_gram_value(row[4] if len(row) > 4 else ""),
                    "fat": parse_gram_value(row[5] if len(row) > 5 else ""),
                    "carbs": parse_gram_value(row[6] if len(row) > 6 else ""),
                    "fiber": parse_gram_value(row[7] if len(row) > 7 else ""),
                    "sodium": parse_gram_value(row[8] if len(row) > 8 else ""),
                    "calories": parse_calories(row[14] if len(row) > 14 else ""),
                    "day_schedule": day_text,
                    "_weekdays": set() if parsed_days is None else set(parsed_days),
                    "_daily": parsed_days is None,
                }
            else:
                item = merged[key]
                if parsed_days is None:
                    item["_daily"] = True
                    item["_weekdays"].clear()
                elif not item["_daily"]:
                    item["_weekdays"].update(parsed_days)
                if day_text and day_text not in (item.get("day_schedule") or ""):
                    item["day_schedule"] = f"{item.get('day_schedule', '')}、{day_text}".strip("、")

    rows = []
    for item in merged.values():
        if item.pop("_daily"):
            item["available_weekdays"] = None
        else:
            days = sorted(item.pop("_weekdays"))
            item["available_weekdays"] = days if days else None
        item.pop("day_schedule", None)
        rows.append(item)
    return rows


def load_xueyi_shaoyuan_excel(xlsx_path: str) -> List[Dict]:
    import openpyxl

    rows = []
    current_loc = ""
    current_canteen = "学一食堂"
    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    ws = wb[wb.sheetnames[0]]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(c for c in row if c):
            continue
        if row[0]:
            current_loc = str(row[0]).strip()
            current_canteen = resolve_xueyi_shaoyuan_canteen(current_loc, current_canteen)
        name = str(row[1]).strip() if row[1] else ""
        if not name or name == "菜品名称":
            continue
        cal_raw = row[5] if len(row) > 5 else 400
        rows.append({
            "canteen": current_canteen,
            "location": current_loc,
            "name": name,
            "ingredients": str(row[2] or "").strip() if len(row) > 2 else "",
            "flavor_text": str(row[3] or "").strip() if len(row) > 3 else "",
            "nutrition_text": str(row[4] or "").strip() if len(row) > 4 else "",
            "calories_raw": cal_raw,
        })
    return rows


def load_songlin_excel(xlsx_path: str) -> List[Dict]:
    import openpyxl

    rows = []
    current_loc = ""
    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    ws = wb[wb.sheetnames[0]]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(c for c in row if c):
            continue
        if row[0]:
            current_loc = str(row[0]).strip()
        name = str(row[1]).strip() if row[1] else ""
        if not name:
            continue
        rows.append({
            "location": current_loc,
            "name": name,
            "ingredients": str(row[2] or "").strip() if len(row) > 2 else "",
            "flavor_text": str(row[3] or "").strip() if len(row) > 3 else "",
            "nutrition_text": str(row[4] or "").strip() if len(row) > 4 else "",
            "calories_raw": row[5] if len(row) > 5 else "",
        })
    return rows


def tongyuan_row_to_dish(
    row: Dict,
    dish_id: str,
    window_numbers: Dict[str, int],
) -> Dict:
    canteen = "佟园"
    window = parse_tongyuan_window(row.get("location", ""))
    if window not in window_numbers:
        window_numbers[window] = len(window_numbers) + 1

    name = row["name"]
    ingredients = row.get("ingredients", "")
    flavor_text = row.get("flavor_text", "")
    calories = row.get("calories") or 400
    nutrition = {
        "calories": int(calories),
        "protein": row.get("protein", 0),
        "fat": row.get("fat", 0),
        "carbs": row.get("carbs", 0),
        "fiber": row.get("fiber", 2),
    }
    price = infer_price_from_context(name, window, calories)
    meta = CANTEEN_META[canteen]

    dish = {
        "id": dish_id,
        "name": name,
        "canteen": canteen,
        "canteen_id": meta["canteen_id"],
        "window": window,
        "window_number": window_numbers[window],
        "floor": 1,
        "price": price,
        "cuisine": infer_cuisine(window, canteen, flavor_text),
        "flavor": parse_flavor_text(flavor_text) or infer_flavor(name, ingredients),
        "cooking": "现做",
        "appearance": 3,
        "portion_size": "M" if calories < 350 else "L",
        "prep_time": 8,
        "image": "",
        "tags": infer_tags(name, window, price),
        "rating": 4.0,
        "rating_count": 0,
        "hours": {"lunch": True, "dinner": True, "late_night": False},
        "related_dishes": [],
        "location_hint": _simple_location_hint(canteen, window),
        "ingredients": ingredients,
        **nutrition,
    }
    if row.get("available_weekdays") is not None:
        dish["available_weekdays"] = row["available_weekdays"]
    if row.get("sodium"):
        dish["sodium_mg"] = row["sodium"]
    return dish


def xueyi_shaoyuan_row_to_dish(
    row: Dict,
    dish_id: str,
    window_numbers: Dict[str, int],
) -> Dict:
    canteen = row["canteen"]
    window = (row.get("location") or "综合").strip()
    if window not in window_numbers:
        window_numbers[window] = len(window_numbers) + 1

    name = row["name"]
    ingredients = row.get("ingredients", "")
    flavor_text = row.get("flavor_text", "")
    calories = parse_calories(row.get("calories_raw", 400))
    nutrition = infer_nutrition(calories)
    price = infer_price_from_context(name, window, calories)
    meta = CANTEEN_META.get(canteen, {"canteen_id": canteen, "cuisine_default": "融合"})
    hint = f"{canteen}，{window}" if canteen not in SIMPLE_LOCATION_CANTEENS else canteen

    dish = {
        "id": dish_id,
        "name": name,
        "canteen": canteen,
        "canteen_id": meta["canteen_id"],
        "window": window,
        "window_number": window_numbers[window],
        "floor": 1,
        "price": price,
        "cuisine": infer_cuisine(window, canteen, flavor_text),
        "flavor": parse_flavor_text(flavor_text) or infer_flavor(name, ingredients),
        "cooking": "现做",
        "appearance": 3,
        "portion_size": "M" if calories < 350 else "L",
        "prep_time": 8,
        "image": "",
        "tags": infer_tags(name, window, price),
        "rating": 4.0,
        "rating_count": 0,
        "hours": {"lunch": True, "dinner": True, "late_night": False},
        "related_dishes": [],
        "location_hint": hint,
        "ingredients": ingredients,
        **nutrition,
    }
    if row.get("nutrition_text"):
        dish["nutrition_notes"] = row["nutrition_text"]
    return dish


def songlin_row_to_dish(
    row: Dict,
    dish_id: str,
    window_numbers: Dict[str, int],
) -> Dict:
    canteen = "松林"
    window = parse_songlin_window(row.get("location", ""))
    if window not in window_numbers:
        window_numbers[window] = len(window_numbers) + 1

    name = row["name"]
    ingredients = row.get("ingredients", "")
    flavor_text = row.get("flavor_text", "")
    calories = parse_calories(row.get("calories_raw", 300))
    nutrition = infer_nutrition(calories)
    price = infer_price_from_context(name, window, calories)
    if any(k in window for k in ("包子", "粥")):
        price = min(price, 8.0) if price > 10 else price
    meta = CANTEEN_META[canteen]

    return {
        "id": dish_id,
        "name": name,
        "canteen": canteen,
        "canteen_id": meta["canteen_id"],
        "window": window,
        "window_number": window_numbers[window],
        "floor": 1,
        "price": price,
        "cuisine": infer_cuisine(window, canteen, flavor_text),
        "flavor": parse_flavor_text(flavor_text) or infer_flavor(name, ingredients),
        "cooking": "现做",
        "appearance": 3,
        "portion_size": "S" if price <= 8 else "M",
        "prep_time": 5,
        "image": "",
        "tags": infer_tags(name, window, price) or ["早餐"],
        "rating": 4.0,
        "rating_count": 0,
        "hours": {"lunch": True, "dinner": True, "late_night": False},
        "related_dishes": [],
        "location_hint": _simple_location_hint(canteen, window),
        "ingredients": ingredients,
        **nutrition,
    }


def _remove_canteens(existing: List[Dict], names: set) -> List[Dict]:
    return [d for d in existing if d.get("canteen") not in names]


def merge_tongyuan_csv(
    dishes_file: str,
    csv_path: str,
    *,
    id_prefix: str = "ty",
    replace_existing: bool = True,
) -> Tuple[int, int]:
    path = Path(dishes_file)
    existing = json.loads(path.read_text(encoding="utf-8")) if path.exists() else []
    if replace_existing:
        existing = _remove_canteens(existing, {"佟园"})

    raw_rows = load_tongyuan_csv(csv_path)
    existing_names = {(d["canteen"], d["name"], d.get("window", "")) for d in existing}
    existing_ids = {d["id"] for d in existing}
    window_numbers: Dict[str, int] = {}
    added = skipped = 0
    idx = 1

    for row in raw_rows:
        window = parse_tongyuan_window(row.get("location", ""))
        key = ("佟园", row["name"], window)
        if key in existing_names:
            skipped += 1
            continue
        dish_id, idx = _next_dish_id(id_prefix, idx, existing_ids)
        dish = tongyuan_row_to_dish(row, dish_id, window_numbers)
        existing.append(dish)
        existing_names.add(key)
        existing_ids.add(dish_id)
        added += 1
        idx += 1

    path.write_text(json.dumps(existing, ensure_ascii=False, indent=2), encoding="utf-8")
    return added, skipped


def merge_xueyi_shaoyuan_excel(
    dishes_file: str,
    xlsx_path: str,
    *,
    id_prefix_xueyi: str = "xy",
    id_prefix_shaoyuan: str = "sy",
    replace_existing: bool = True,
) -> Dict[str, Tuple[int, int]]:
    path = Path(dishes_file)
    existing = json.loads(path.read_text(encoding="utf-8")) if path.exists() else []
    if replace_existing:
        existing = _remove_canteens(existing, {"学一食堂", "勺园"})

    raw_rows = load_xueyi_shaoyuan_excel(xlsx_path)
    existing_names = {(d["canteen"], d["name"], d.get("window", "")) for d in existing}
    existing_ids = {d["id"] for d in existing}
    window_numbers: Dict[str, int] = {}
    stats = {"学一食堂": [0, 0], "勺园": [0, 0]}
    idx_map = {"学一食堂": 1, "勺园": 1}
    prefix_map = {"学一食堂": id_prefix_xueyi, "勺园": id_prefix_shaoyuan}

    for row in raw_rows:
        canteen = row["canteen"]
        name = row["name"]
        window = (row.get("location") or "综合").strip()
        if (canteen, name, window) in existing_names:
            stats[canteen][1] += 1
            continue
        prefix = prefix_map[canteen]
        idx = idx_map[canteen]
        dish_id, idx = _next_dish_id(prefix, idx, existing_ids)
        dish = xueyi_shaoyuan_row_to_dish(row, dish_id, window_numbers)
        existing.append(dish)
        existing_names.add((canteen, name, window))
        existing_ids.add(dish_id)
        stats[canteen][0] += 1
        idx_map[canteen] = idx + 1

    path.write_text(json.dumps(existing, ensure_ascii=False, indent=2), encoding="utf-8")
    return {k: tuple(v) for k, v in stats.items()}


def merge_songlin_excel(
    dishes_file: str,
    xlsx_path: str,
    *,
    id_prefix: str = "sl",
    replace_existing: bool = True,
) -> Tuple[int, int]:
    path = Path(dishes_file)
    existing = json.loads(path.read_text(encoding="utf-8")) if path.exists() else []
    if replace_existing:
        existing = _remove_canteens(existing, {"松林"})

    raw_rows = load_songlin_excel(xlsx_path)
    existing_names = {(d["canteen"], d["name"]) for d in existing}
    existing_ids = {d["id"] for d in existing}
    window_numbers: Dict[str, int] = {}
    added = skipped = 0
    idx = 1

    for row in raw_rows:
        name = row["name"]
        if ("松林", name) in existing_names:
            skipped += 1
            continue
        dish_id, idx = _next_dish_id(id_prefix, idx, existing_ids)
        dish = songlin_row_to_dish(row, dish_id, window_numbers)
        existing.append(dish)
        existing_names.add(("松林", name))
        existing_ids.add(dish_id)
        added += 1
        idx += 1

    path.write_text(json.dumps(existing, ensure_ascii=False, indent=2), encoding="utf-8")
    return added, skipped


def merge_yannan_and_jiayuan(
    dishes_file: str,
    yannan_json: str,
    jiayuan_json: str,
) -> Dict[str, Tuple[int, int]]:
    """一次性导入燕南地下 + 家园三层"""
    results = {}
    results["燕南美食"] = merge_json_canteen_file(
        dishes_file,
        yannan_json,
        target_canteen="燕南美食",
        floor=0,
        location_prefix="燕南地下",
        id_prefix="ynd",
        replace_existing=True,
    )
    # 燕南已清空；家园需在第二次合并时保留刚导入的燕南
    results["家园食堂"] = merge_json_canteen_file(
        dishes_file,
        jiayuan_json,
        target_canteen="家园食堂",
        floor=3,
        location_prefix="家园三层",
        id_prefix="jy3",
        replace_existing=True,
    )
    return results
